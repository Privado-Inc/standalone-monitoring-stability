import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import functools
from math import floor
import builder
import config
from utils.DiffCache import DiffCache, TimeDiffCache
from utils.helpers import print_timestamp


def add_missing_emoji(missing_value):
    return ":rotating_light:" if missing_value > 0 else ""

def create_new_excel_for_file(location, first_file, second_file):
    wb = openpyxl.Workbook()
    # Delete the default sheet
    wb.remove(wb.active)
    # Create new Sheet
    wb.create_sheet('scan-status')
    wb.create_sheet('source-&-sink-report')
    wb.create_sheet('flow-report')
    wb.save(location)


def create_new_excel(location, base_branch_name, head_branch_name):
    wb = openpyxl.Workbook()
    # Delete the default sheet
    wb.remove(wb.active)
    # Create new Sheet
    wb.create_sheet('summary')
    wb.create_sheet('scan-status')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-source-&-sink-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-flow-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-unique-flow-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-collections-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-performance-report')
    wb.save(location)


def write_source_sink_data(workbook_location, worksheet_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_path_data(workbook_location, worksheet_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_performance_data(workbook_location, worksheet_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_scan_status_report_for_file(workbook_location, base_branch_name, head_branch_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    worksheet = workbook['scan-status']

    for line in report:
        worksheet.append(line)

    workbook.save(workbook_location)


def write_scan_status_report(workbook_location, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    worksheet = workbook['scan-status']

    worksheet.append(
        ["Repo", "Branch", "language" ,"scan status", "scan error", "comparison status", "comparison error", "unique flow count",
         "scan time (ms)", "CPG size"])

    for repo in report.keys():
        repo_info = report[repo]
        worksheet.append([repo, config.BASE_CORE_BRANCH_KEY,
                          report[repo]['language'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['scan_status'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['scan_error_message'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['comparison_status'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['comparison_error_message'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['unique_flows'],
                          # repo_info[config.BASE_CORE_BRANCH_KEY]['code_scan_time'],
                          repo_info[config.BASE_CORE_BRANCH_KEY]['binary_file_size']])

        worksheet.append([repo, config.HEAD_CORE_BRANCH_KEY,
                          report[repo]['language'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['scan_status'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['scan_error_message'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['comparison_status'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['comparison_error_message'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['unique_flows'],
                          # repo_info[config.HEAD_CORE_BRANCH_KEY]['code_scan_time'],
                          repo_info[config.HEAD_CORE_BRANCH_KEY]['binary_file_size']])

    workbook.save(workbook_location)


def write_summary_data(workbook_location, report, data_elements, collections ,flow_report):
    workbook = openpyxl.load_workbook(filename=workbook_location)
    worksheet = workbook['summary']

    worksheet.append(["Repo", "language", "scan status", f"{config.BASE_SHEET_BRANCH_NAME} Scan status (ms)", f"{config.HEAD_SHEET_BRANCH_NAME} scan time (ms)",
                      "scan time diff (ms)", "Reachable by flow diff (ms)", f"{config.BASE_SHEET_BRANCH_NAME} unique flows",
                      f"{config.HEAD_SHEET_BRANCH_NAME} unique flows", "unique flows diff",
                      f"{config.BASE_SHEET_BRANCH_NAME} No of data elements",
                      f"{config.HEAD_SHEET_BRANCH_NAME} No of data elements", "Data element diff",
                      f"Missing sinks in {config.HEAD_SHEET_BRANCH_NAME}",
                      "No of 100% missing source to sink combinations", f"Total collections in {config.BASE_SHEET_BRANCH_NAME}", f"Total collections in {config.HEAD_SHEET_BRANCH_NAME}", f"Unique collection diff"])

    scan_time_positive = 0
    scan_time_positive_average = 0
    scan_time_negative_average = 0

    reachable_by_flow_time_positive = 0
    reachable_by_flow_time_positive_average = 0
    reachable_by_flow_time_negative_average = 0

    matching_flows = 0
    more_flows = 0
    less_flows = 0

    more_sources = 0
    less_sources = 0
    matching_sources = 0

    missing_sink_repo_count = 0
    total_missing_sinks = 0

    matching_collections = 0
    less_collections = 0
    more_collections = 0

    additional_not_zero = len(list(filter(lambda x: x['additional'] > 0, flow_report.values())))
    try:
        additional_average = functools.reduce(lambda a, x: a + x['additional'], flow_report.values(), 0) / additional_not_zero
    except Exception as e:
        print(e)
        additional_average = 0

    missing_not_zero = len(list(filter(lambda x: x['missing'] > 0, flow_report.values())))
    try:
        missing_average = functools.reduce(lambda a, x: a + x['missing'], flow_report.values(), 0) / missing_not_zero
    except Exception as e:
        print(e)
        missing_average = 0

    matching_flow_repo_count = len(list(filter(lambda x: x['matching_flows'], flow_report.values())))
    hundred_percent_missing = len(list(filter(lambda x: x['hundred_missing'] > 0, flow_report.values())))    
    scan_failure = ""
    num_repos_failed = 0

    for repo in report.keys():
        try:
            language = report[repo]['language']
            scan_status = 'done' if report[repo][config.HEAD_CORE_BRANCH_KEY]['comparison_error_message'] == '--' and report[repo][config.BASE_CORE_BRANCH_KEY]['comparison_error_message'] == '--' else 'failed'
            scan_failure += f"\tScan failed for {repo} - language: {language}" if scan_status == "failed" else ""
            num_repos_failed += 1 if scan_status == "failed" else 0
            head_scan_time = report[repo][config.HEAD_CORE_BRANCH_KEY]['code_scan_time'].split()[0]
            base_scan_time = report[repo][config.BASE_CORE_BRANCH_KEY]['code_scan_time'].split()[0]
            scan_time_diff = '--' if base_scan_time == '--' or head_scan_time == '--' else int(head_scan_time) - int(base_scan_time)
            unique_source_diff = '--' if data_elements[repo][config.BASE_CORE_BRANCH_KEY] == '--' or data_elements[repo][config.HEAD_CORE_BRANCH_KEY] == '--' else int(data_elements[repo][config.HEAD_CORE_BRANCH_KEY]) - int(data_elements[repo][config.BASE_CORE_BRANCH_KEY])
            unique_flow_diff = '--' if report[repo][config.BASE_CORE_BRANCH_KEY]['unique_flows'] == '--' or report[repo][config.HEAD_CORE_BRANCH_KEY]['unique_flows'] == '--' else int(report[repo][config.HEAD_CORE_BRANCH_KEY]['unique_flows']) - int(report[repo][config.BASE_CORE_BRANCH_KEY]['unique_flows'])
            reachable_flow_time_diff = '--' if report[repo][config.BASE_CORE_BRANCH_KEY]['reachable_flow_time'] == '--' or report[repo][config.HEAD_CORE_BRANCH_KEY]['reachable_flow_time'] == '--' else int(report[repo][config.HEAD_CORE_BRANCH_KEY]['reachable_flow_time']) - int(report[repo][config.BASE_CORE_BRANCH_KEY]['reachable_flow_time'])
            number_hundred_missing_for_repo = flow_report[repo]['hundred_missing']
                        


            unique_collections_diff = '--' if collections[repo][config.BASE_CORE_BRANCH_KEY] == '--' or collections[repo][config.HEAD_CORE_BRANCH_KEY] == '--' else int(collections[repo][config.HEAD_CORE_BRANCH_KEY]) - int(collections[repo][config.BASE_CORE_BRANCH_KEY])
            if scan_time_diff != '--':
                if scan_time_diff > 0: # Head branch took more time
                    scan_time_positive += 1
                    scan_time_positive_average += scan_time_diff
                else:
                    scan_time_negative_average += (scan_time_diff*-1)

            if unique_flow_diff != '--':
                if unique_flow_diff > 0:
                    more_flows += 1
                elif unique_flow_diff < 0:
                    less_flows += 1
                else:
                    matching_flows += 1

            if unique_source_diff != '--':
                if unique_source_diff > 0:
                    more_sources += 1
                elif unique_source_diff < 0:
                    less_sources += 1
                else:
                    matching_sources += 1

            if unique_collections_diff != '--':
                if unique_collections_diff > 0:
                    more_collections += 1
                elif unique_collections_diff < 0:
                    less_collections += 1
                else:
                    matching_collections += 1
        
            
            if reachable_flow_time_diff != '--':
                # Head branch took more time
                if reachable_flow_time_diff > 0:
                    reachable_by_flow_time_positive += 1
                    reachable_by_flow_time_positive_average += reachable_flow_time_diff
                else:
                    reachable_by_flow_time_negative_average += (reachable_flow_time_diff*-1)

            if int(report[repo]['missing_sink']) > 0:
                missing_sink_repo_count += 1
                total_missing_sinks += int(report[repo]['missing_sink'])

            worksheet.append([repo, language, scan_status, base_scan_time, head_scan_time, scan_time_diff,
                              reachable_flow_time_diff,
                              report[repo][config.BASE_CORE_BRANCH_KEY]['unique_flows'],
                              report[repo][config.HEAD_CORE_BRANCH_KEY]['unique_flows'], unique_flow_diff,
                              data_elements[repo][config.BASE_CORE_BRANCH_KEY],
                              data_elements[repo][config.HEAD_CORE_BRANCH_KEY], unique_source_diff,
                              report[repo]['missing_sink'],
                              number_hundred_missing_for_repo, collections[repo][config.BASE_CORE_BRANCH_KEY], collections[repo][config.HEAD_CORE_BRANCH_KEY], unique_collections_diff])

        except Exception as e:
            print_timestamp(f'Scan failed for repo {repo} : {str(e)}')
            worksheet.append([repo, language, scan_status, "--", "--", "--",
                              "--",
                              "--",
                              "--",
                              "--",
                              "--",
                              "--",
                              "--",
                              "--",
                              "--"])

    # cannot divide by zero
    scan_time_positive_average = TimeDiffCache["scan_time"]["more"][1] / TimeDiffCache["scan_time"]["more"][0] if TimeDiffCache["scan_time"]["more"][0] > 0 else 0 # Average of more time repos
    scan_time_negative_average = TimeDiffCache["scan_time"]["less"][1] / TimeDiffCache["scan_time"]["less"][0] if TimeDiffCache["scan_time"]["less"][0] > 0 else 0 # Average of less time repos
    
    reachable_by_flow_time_positive_average = TimeDiffCache["reachable_by_flow_time"]["more"][1] / TimeDiffCache["reachable_by_flow_time"]["more"][0] if TimeDiffCache["reachable_by_flow_time"]["more"][0] > 0 else 0 # Average of more time repos
    reachable_by_flow_time_negative_average = TimeDiffCache["reachable_by_flow_time"]["less"][1] / TimeDiffCache["reachable_by_flow_time"]["less"][0] if TimeDiffCache["reachable_by_flow_time"]["less"][0] > 0 else 0 # Average of less time repos

    missing_sink_average = total_missing_sinks // missing_sink_repo_count if missing_sink_repo_count > 0 else 0
    
    write_slack_summary(f'''
        A. Repository scan failure report
        Scan for {num_repos_failed} out of {len(list(report.keys()))} repositories failed. {":rotating_light:" if num_repos_failed > 0 else ""}
{scan_failure if len(scan_failure) > 0 else ""} 
        B. Scan time difference
        {TimeDiffCache["scan_time"]["more"][0]} repositories took on an average {floor(scan_time_positive_average)} ms more. 
        {TimeDiffCache["scan_time"]["less"][0]} repositories took on an average {floor(scan_time_negative_average)} ms less.
        
        C. Reachable by Flow time difference.
        {TimeDiffCache["reachable_by_flow_time"]["more"][0]} repositories took on an average {floor(reachable_by_flow_time_positive_average)} ms more. 
        {TimeDiffCache["reachable_by_flow_time"]["less"][0]} repositories took on an average {floor(reachable_by_flow_time_negative_average)} ms less.
        
        D. Reachable by Flow count difference.
        {DiffCache["reachable_by_flow_count"]["matching"]} repositories have exactly matching flows.
        {DiffCache["reachable_by_flow_count"]["missing"]} repositories have missing flows. {add_missing_emoji(DiffCache["reachable_by_flow_count"]["missing"])}
        {DiffCache["reachable_by_flow_count"]["additional"]} repositories have additional flows.
        
        E. Unique data elements difference.
        {DiffCache["sources"]["matching"]} repositories have exactly matching elements.
        {DiffCache["sources"]["missing"]} repositories have missing data elements. {add_missing_emoji(less_sources)}
        {DiffCache["sources"]["additional"]} repositories have additional elements.

        F. Missing sinks.
        {DiffCache["sinks"]["missing"]} repositories have missing sinks. {add_missing_emoji(DiffCache["sinks"]["missing"])}

        G. Source to Sink Flow data
        {hundred_percent_missing} repositories have hundred percent missing flows. {add_missing_emoji(hundred_percent_missing)} {add_missing_emoji(hundred_percent_missing)}
        {DiffCache["dataflows"]["matching"]} repositories have matching flows.
        {DiffCache["dataflows"]["additional"]} repositories have on an average {floor(additional_average)} additional flows.
        {DiffCache["dataflows"]["missing"]} repositories have on an average {floor(missing_average)} missing flows. {add_missing_emoji(missing_not_zero)}
        
        H. Collection Summary
        {DiffCache["collections"]["matching"]} repositories have exactly matching collections.
        {DiffCache["collections"]["missing"]} repositories have missing collections. {add_missing_emoji(DiffCache["collections"]["missing"])}
        {DiffCache["collections"]["additional"]} repositories have additional collections.
    ''')
    
    highlight_summary_cell(worksheet)
    workbook.save(workbook_location)


def highlight_summary_cell(worksheet):

    for col in ['F', 'G', 'J', 'M', 'N', 'R']:
        for row in range(2, len(worksheet[col]) + 1):
            try:
                if int(worksheet[f'{col}{row}'].value) < 0:
                    worksheet[f'{col}{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
                    worksheet[f'{col}{row}'].font = Font(color='FFFFFF')
                else:
                    worksheet[f'{col}{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    worksheet[f'{col}{row}'].font = Font(color='FFFFFF')
            except:
                pass
                

def write_slack_summary(statement):
    with open(f"{os.getcwd()}/slack_summary.txt", "a") as slack_summary:
        slack_summary.writelines(statement)


def write_to_action_result(content):
    file_path = f"{os.getcwd()}/action_result.txt"
    if not os.path.exists(file_path):
        with open(file_path, "w") as action_result:
            action_result.write(f"{content}\n")
    else:
        with open(file_path, "a") as action_result:
            action_result.write(f"{content}\n")
